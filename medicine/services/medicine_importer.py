from decimal import Decimal

from django.core.exceptions import ImproperlyConfigured
from django.db import transaction

from medicine.models import Medicine, MedicineVariant

def import_medicines_from_excel(file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImproperlyConfigured(
            "Excel import requires openpyxl. Install project requirements and try again."
        ) from exc

    workbook = load_workbook(file)
    sheet = workbook.active

    success_count = 0 
    errors = [] 

    with transaction.atomic():
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),start=2):

            try:

                if not any(row):
                    continue
                    
                medicine_name = str(row[0] or "").strip().title()
                short_name = str(row[1] or "").strip().upper()
                medicine_type = str(row[2] or "").strip().lower()
                company = str(row[3] or "").strip()

                power = str(row[4] or "").strip()

                if not medicine_name:
                    raise Exception("Medicine name missing")

                if not short_name:
                    raise Exception("Short name missing")

                if not power:
                    raise Exception("Power missing")

                valid_types = {choice[0] for choice in Medicine.MEDICINE_TYPE}
                if medicine_type not in valid_types:
                    raise Exception(f"Invalid medicine type: {medicine_type}")

                unit_per_strip = int(row[5] or 1)

                cost_price = Decimal(str(row[6] or 0))
                selling_price = Decimal(str(row[7] or 0))

                stock_strips = int(row[8] or 0)
                
                stock_units = stock_strips * unit_per_strip

                medicine, created = Medicine.objects.get_or_create(
                    short_name = short_name,
                    defaults = {
                        "name":medicine_name,
                        "medicine_type":medicine_type,
                        "company":company,
                        "is_active": True,
                    }
                )

                variant = MedicineVariant.objects.filter(
                    medicine = medicine,
                    power = power
                ).first()

                if variant:
                    variant.stock += stock_units 
                    variant.cost_price = cost_price
                    variant.selling_price = selling_price
                    variant.unit_per_strip = unit_per_strip

                    variant.save()
                else:
                    MedicineVariant.objects.create(
                        medicine = medicine,
                        power = power,
                        cost_price = cost_price,
                        selling_price = selling_price,
                        stock = stock_units,
                        unit = 'piece',
                        unit_per_strip = unit_per_strip
                    )

                success_count += 1 
            except Exception as e :
                errors.append(
                    f"Row {row_number}: {str(e)}"
                )

        
    return {
        "success_count": success_count,
        "errors": errors,
    }
